#!/usr/bin/env python3
"""
发票业务服务层

该模块将原本耦合在 UI 层的发票业务逻辑（批量重命名、导出 Excel、批量字段提取）
抽象为独立的 Python 服务，不依赖 Qt，可供 GUI 和 MCP Server 共用。
"""

import logging
import os
import re
from datetime import datetime
from typing import List, Dict, Any, Optional, Callable

from openpyxl import Workbook
from openpyxl.styles import PatternFill


# 模块级日志器
logger = logging.getLogger(__name__)


class InvoiceService:
    """
    发票业务服务类

    提供与界面无关的发票处理功能，包括批量字段提取、批量重命名、导出 Excel 等。

    Attributes:
        pdf_handler: PDFHandler 实例，用于执行 PDF 合并和字段提取
        INVALID_CHARS: 文件名非法字符正则
    """

    INVALID_CHARS = r'[<>:"/\\|?*]'

    # 重命名规则中可用的字段映射（中文占位符 -> info 字典键）
    RENAME_FIELD_MAPPING = {
        '发票类型': 'invoice_type',
        '商品类型': 'product_type',
        '开票日期': 'invoice_date',
        '发票号码': 'invoice_code',
        '买方名字': 'buyer_name',
        '销方名字': 'seller_name',
        '金额': 'amount',
    }

    def __init__(self, pdf_handler):
        """
        初始化发票业务服务

        参数:
            pdf_handler: PDFHandler 实例，提供 merge_pdfs 和 extract 系列方法
        """
        self.pdf_handler = pdf_handler

    def batch_extract(self, pdf_paths: List[str],
                      on_progress: Optional[Callable[[int, int], None]] = None) -> List[Dict[str, Any]]:
        """
        批量提取发票字段信息

        功能描述:
            依次调用 PDFHandler.extract_all_invoice_info 提取每张发票的字段。

        参数:
            pdf_paths: PDF 文件路径列表
            on_progress: 进度回调函数，签名为 on_progress(current, total)

        返回值:
            List[Dict[str, Any]]: 每个元素包含 'path' 和 'info' 的字典列表
        """
        results = []
        total = len(pdf_paths)

        for idx, pdf_path in enumerate(pdf_paths, start=1):
            try:
                info = self.pdf_handler.extract_all_invoice_info(pdf_path)
                results.append({
                    'path': pdf_path,
                    'info': info,
                })
            except Exception as e:
                logger.exception("批量提取失败: %s", pdf_path)
                results.append({
                    'path': pdf_path,
                    'info': None,
                })

            if on_progress:
                on_progress(idx, total)

        return results

    def batch_rename(self, pdf_paths: List[str], rule: str,
                     on_progress: Optional[Callable[[int, int], None]] = None,
                     dry_run: bool = False) -> Dict[str, Any]:
        """
        根据规则批量重命名发票 PDF 文件

        功能描述:
            提取每个 PDF 的发票字段，应用规则生成新文件名，处理文件名冲突后执行或预览重命名。

        参数:
            pdf_paths: 需要重命名的 PDF 文件路径列表
            rule: 重命名规则字符串，支持 {发票类型}、{开票日期} 等占位符
            on_progress: 进度回调函数，签名为 on_progress(current, total)
            dry_run: 是否为预览模式；为 True 时不实际重命名文件，仅返回预计映射

        返回值:
            Dict[str, Any]: 重命名结果，包含以下字段
                - renamed_count: 成功重命名（或预览中将会重命名）的文件数
                - failed_count: 失败的文件数
                - unrecognized_files: 无法识别（info 为 None）的文件名列表
                - renamed_map: 旧路径 -> 新路径 的映射字典
        """
        renamed_count = 0
        failed_count = 0
        unrecognized_files = []
        renamed_map = {}
        total = len(pdf_paths)

        # 预览模式下用于模拟文件名冲突的已占用名称集合
        used_names = set(os.path.basename(p) for p in pdf_paths) if dry_run else None

        for idx, file_path in enumerate(pdf_paths, start=1):
            try:
                file_info = self.pdf_handler.extract_all_invoice_info(file_path)

                if file_info is None:
                    unrecognized_files.append(os.path.basename(file_path))
                    continue

                new_name = self.apply_rule(rule, file_info)
                if not new_name:
                    failed_count += 1
                    continue

                _, ext = os.path.splitext(file_path)
                new_name_with_ext = new_name + ext

                # 处理文件名冲突
                dir_path = os.path.dirname(file_path)
                final_name = new_name_with_ext
                counter = 1

                if dry_run:
                    while final_name in used_names:
                        if final_name == os.path.basename(file_path):
                            break
                        base_name = f"{new_name}({counter})"
                        final_name = base_name + ext
                        counter += 1
                else:
                    while os.path.exists(os.path.join(dir_path, final_name)):
                        if final_name == os.path.basename(file_path):
                            break
                        base_name = f"{new_name}({counter})"
                        final_name = base_name + ext
                        counter += 1

                if final_name != os.path.basename(file_path):
                    new_path = os.path.join(dir_path, final_name)
                    if dry_run:
                        used_names.add(final_name)
                        used_names.discard(os.path.basename(file_path))
                    else:
                        os.rename(file_path, new_path)
                    renamed_count += 1
                    renamed_map[file_path] = new_path

            except Exception as e:
                logger.exception("重命名文件失败: %s", file_path)
                failed_count += 1

            if on_progress:
                on_progress(idx, total)

        return {
            'renamed_count': renamed_count,
            'failed_count': failed_count,
            'unrecognized_files': unrecognized_files,
            'renamed_map': renamed_map,
        }

    def export_to_excel(self, pdf_infos: List[Dict[str, Any]], output_path: str,
                        duplicate_codes: Optional[set] = None) -> bool:
        """
        将发票信息列表导出为 Excel 文件

        功能描述:
            使用 openpyxl 创建 Excel 工作簿，写入发票列表、标记重复发票、添加总计行。

        参数:
            pdf_infos: 发票信息字典列表，每个字典需包含 name、amount、tax_amount、
                       invoice_date、invoice_code、path、mod_time、size 等字段
            output_path: 输出 Excel 文件路径
            duplicate_codes: 重复发票号码集合，用于高亮标记，可选

        返回值:
            bool: 导出是否成功
        """
        if duplicate_codes is None:
            duplicate_codes = set()

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "发票列表"

            headers = ["发票代码", "文件名", "金额", "税额", "发票日期", "文件路径", "修改时间", "大小"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = cell.font.copy(bold=True)

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            total_amount = 0.0
            total_tax = 0.0

            for row_idx, file_info in enumerate(pdf_infos, 2):
                invoice_code = file_info.get('invoice_code', '')
                amount = file_info.get('amount', 0.0)
                tax_amount = file_info.get('tax_amount', 0.0)

                if isinstance(amount, (int, float)):
                    total_amount += amount
                if isinstance(tax_amount, (int, float)):
                    total_tax += tax_amount

                is_duplicate = invoice_code in duplicate_codes and invoice_code != ""

                ws.cell(row=row_idx, column=1, value=invoice_code)
                ws.cell(row=row_idx, column=2, value=file_info.get('name', ''))
                ws.cell(row=row_idx, column=3, value=amount)
                ws.cell(row=row_idx, column=4, value=tax_amount)
                ws.cell(row=row_idx, column=5, value=file_info.get('invoice_date', ''))
                ws.cell(row=row_idx, column=6, value=file_info.get('path', ''))
                ws.cell(row=row_idx, column=7, value=file_info.get('mod_time', ''))
                ws.cell(row=row_idx, column=8, value=file_info.get('size', ''))

                if is_duplicate:
                    for col in range(1, 9):
                        ws.cell(row=row_idx, column=col).fill = yellow_fill

            # 总计行
            total_row = len(pdf_infos) + 2
            ws.cell(row=total_row, column=1, value="")
            ws.cell(row=total_row, column=2, value="总计")
            ws.cell(row=total_row, column=3, value=total_amount)
            ws.cell(row=total_row, column=4, value=total_tax)
            ws.cell(row=total_row, column=5, value="")
            ws.cell(row=total_row, column=6, value="")
            ws.cell(row=total_row, column=7, value="")
            ws.cell(row=total_row, column=8, value="")

            for col in range(1, 9):
                ws.cell(row=total_row, column=col).font = ws.cell(row=total_row, column=col).font.copy(bold=True)

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 12

            wb.save(output_path)
            return True

        except Exception as e:
            logger.exception("导出 Excel 失败: %s", output_path)
            return False

    @staticmethod
    def apply_rule(rule: str, file_info: Dict[str, Any]) -> Optional[str]:
        """
        应用重命名规则生成新文件名

        功能描述:
            将规则字符串中的中文占位符替换为 file_info 中对应的字段值，并清理非法字符。

        参数:
            rule: 重命名规则字符串，如 "{买方名字}-{开票日期}-{商品类型}"
            file_info: 发票信息字典，包含 invoice_type、invoice_date 等字段

        返回值:
            Optional[str]: 清理后的文件名（不含扩展名），规则为空时返回 None
        """
        if not rule:
            return None

        result = rule

        for field_name, field_key in InvoiceService.RENAME_FIELD_MAPPING.items():
            placeholder = f"{{{field_name}}}"
            if placeholder in result:
                value = file_info.get(field_key, '')
                if field_key == 'amount' and isinstance(value, (int, float)):
                    value = f"{value:.2f}"
                result = result.replace(placeholder, str(value))

        result = InvoiceService.sanitize_filename(result)
        return result

    @staticmethod
    def sanitize_filename(filename: str) -> str:
        """
        清理文件名中的非法字符

        功能描述:
            将 Windows 文件名非法字符替换为下划线，并移除首尾的空格和点。

        参数:
            filename: 原始文件名

        返回值:
            str: 清理后的文件名，若为空则返回 "未命名"
        """
        sanitized = re.sub(InvoiceService.INVALID_CHARS, '_', filename)
        sanitized = sanitized.strip('. ')
        if not sanitized:
            sanitized = "未命名"
        return sanitized

    @staticmethod
    def get_available_fields() -> List[Dict[str, str]]:
        """
        获取重命名规则中可用的字段列表

        返回值:
            List[Dict[str, str]]: 每个元素包含 display（显示名）和 key（字典键）
        """
        return [
            {"display": display, "key": key}
            for display, key in InvoiceService.RENAME_FIELD_MAPPING.items()
        ]
