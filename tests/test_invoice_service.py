#!/usr/bin/env python3
"""
InvoiceService 单元测试

不依赖真实 PDF 文件，使用 unittest.mock 伪造 PDFHandler 行为。
"""

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'code'))

from unittest.mock import Mock
from core.invoice_service import InvoiceService


class TestInvoiceServiceStatic(unittest.TestCase):
    """测试 InvoiceService 的静态工具方法"""

    def test_sanitize_filename(self):
        """测试非法字符清理"""
        self.assertEqual(InvoiceService.sanitize_filename('a/b:c.txt'), 'a_b_c.txt')
        self.assertEqual(InvoiceService.sanitize_filename('   '), '未命名')
        self.assertEqual(InvoiceService.sanitize_filename('. '), '未命名')

    def test_apply_rule(self):
        """测试重命名规则替换"""
        info = {
            'invoice_type': '增值税发票',
            'invoice_date': '2024-01-01',
            'amount': 100.5,
            'invoice_code': '123',
            'buyer_name': 'A公司',
            'seller_name': 'B公司',
            'product_type': '办公用品',
        }
        self.assertEqual(
            InvoiceService.apply_rule('{发票类型}-{开票日期}', info),
            '增值税发票-2024-01-01'
        )
        self.assertEqual(InvoiceService.apply_rule('{金额}', info), '100.50')
        self.assertIsNone(InvoiceService.apply_rule('', info))

    def test_get_available_fields(self):
        """测试可用字段列表"""
        fields = InvoiceService.get_available_fields()
        self.assertIsInstance(fields, list)
        self.assertTrue(any(f['key'] == 'amount' for f in fields))


class TestInvoiceServiceBatch(unittest.TestCase):
    """测试需要 PDFHandler 的批量方法"""

    def test_batch_extract(self):
        """测试批量提取失败时返回 None 信息"""
        pdf_handler = Mock()
        pdf_handler.extract_all_invoice_info = Mock(side_effect=[
            {'amount': 1},
            Exception('识别失败'),
        ])
        service = InvoiceService(pdf_handler)
        result = service.batch_extract(['a.pdf', 'b.pdf'])
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]['info']['amount'], 1)
        self.assertIsNone(result[1]['info'])

    def test_batch_rename(self):
        """测试实际重命名"""
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = []
            file_infos = [
                {
                    'invoice_type': '普票',
                    'invoice_date': '2024-01-01',
                    'amount': 10,
                    'invoice_code': '001',
                    'buyer_name': 'A',
                    'seller_name': 'B',
                    'product_type': 'X',
                },
                {
                    'invoice_type': '专票',
                    'invoice_date': '2024-01-02',
                    'amount': 20,
                    'invoice_code': '002',
                    'buyer_name': 'C',
                    'seller_name': 'D',
                    'product_type': 'Y',
                },
            ]
            for idx in range(2):
                p = os.path.join(tmpdir, f'{idx}.pdf')
                with open(p, 'w') as f:
                    pass
                paths.append(p)

            pdf_handler = Mock()
            pdf_handler.extract_all_invoice_info = Mock(side_effect=file_infos)
            service = InvoiceService(pdf_handler)
            result = service.batch_rename(paths, '{发票类型}-{开票日期}')

            self.assertEqual(result['renamed_count'], 2)
            self.assertTrue(os.path.exists(os.path.join(tmpdir, '普票-2024-01-01.pdf')))
            self.assertTrue(os.path.exists(os.path.join(tmpdir, '专票-2024-01-02.pdf')))

    def test_batch_rename_dry_run(self):
        """测试预览模式不修改文件系统"""
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = []
            for name in ['a.pdf', 'b.pdf']:
                p = os.path.join(tmpdir, name)
                with open(p, 'w') as f:
                    pass
                paths.append(p)

            pdf_handler = Mock()
            pdf_handler.extract_all_invoice_info = Mock(return_value={
                'invoice_type': '普票',
                'invoice_date': '2024-01-01',
                'amount': 10,
                'invoice_code': '001',
                'buyer_name': 'A',
                'seller_name': 'B',
                'product_type': 'X',
            })
            service = InvoiceService(pdf_handler)
            result = service.batch_rename(paths, '{发票类型}-{开票日期}', dry_run=True)

            self.assertEqual(result['renamed_count'], 2)
            # 预览模式下文件应保持原名称
            self.assertTrue(os.path.exists(os.path.join(tmpdir, 'a.pdf')))
            self.assertTrue(os.path.exists(os.path.join(tmpdir, 'b.pdf')))

    def test_export_to_excel(self):
        """测试导出 Excel"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'out.xlsx')
            pdf_infos = [
                {
                    'name': 'a.pdf',
                    'amount': 10,
                    'tax_amount': 1,
                    'invoice_date': '2024-01-01',
                    'invoice_code': '001',
                    'path': '/a.pdf',
                    'mod_time': '',
                    'size': '',
                }
            ]
            service = InvoiceService(Mock())
            self.assertTrue(service.export_to_excel(pdf_infos, output_path))
            self.assertTrue(os.path.exists(output_path))

    def test_batch_rename_conflict(self):
        """测试文件名冲突时自动加序号"""
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = []
            file_infos = [
                {
                    'invoice_type': '普票',
                    'invoice_date': '2024-01-01',
                    'amount': 10,
                    'invoice_code': '001',
                    'buyer_name': 'A',
                    'seller_name': 'B',
                    'product_type': 'X',
                },
                {
                    'invoice_type': '普票',
                    'invoice_date': '2024-01-01',
                    'amount': 20,
                    'invoice_code': '002',
                    'buyer_name': 'C',
                    'seller_name': 'D',
                    'product_type': 'Y',
                },
            ]
            for idx in range(2):
                p = os.path.join(tmpdir, f'{idx}.pdf')
                with open(p, 'w') as f:
                    pass
                paths.append(p)

            pdf_handler = Mock()
            pdf_handler.extract_all_invoice_info = Mock(side_effect=file_infos)
            service = InvoiceService(pdf_handler)
            result = service.batch_rename(paths, '{发票类型}-{开票日期}')

            self.assertEqual(result['renamed_count'], 2)
            self.assertTrue(os.path.exists(os.path.join(tmpdir, '普票-2024-01-01.pdf')))
            self.assertTrue(os.path.exists(os.path.join(tmpdir, '普票-2024-01-01(1).pdf')))

    def test_batch_rename_progress(self):
        """测试进度回调被调用 total 次"""
        pdf_handler = Mock()
        pdf_handler.extract_all_invoice_info = Mock(return_value={
            'invoice_type': '普票',
            'invoice_date': '2024-01-01',
            'amount': 10,
            'invoice_code': '001',
            'buyer_name': 'A',
            'seller_name': 'B',
            'product_type': 'X',
        })
        service = InvoiceService(pdf_handler)

        progress_calls = []
        def on_progress(current, total):
            progress_calls.append((current, total))

        with tempfile.TemporaryDirectory() as tmpdir:
            paths = [os.path.join(tmpdir, f'{i}.pdf') for i in range(3)]
            for p in paths:
                with open(p, 'w') as f:
                    pass

            service.batch_rename(paths, '{发票类型}', on_progress=on_progress)

        self.assertEqual(len(progress_calls), 3)
        self.assertEqual(progress_calls[-1], (3, 3))

    def test_export_to_excel_duplicate_highlight(self):
        """测试重复发票代码行被高亮"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'out.xlsx')
            pdf_infos = [
                {
                    'name': 'a.pdf',
                    'amount': 10,
                    'tax_amount': 1,
                    'invoice_date': '2024-01-01',
                    'invoice_code': 'DUP001',
                    'path': '/a.pdf',
                    'mod_time': '',
                    'size': '',
                },
                {
                    'name': 'b.pdf',
                    'amount': 20,
                    'tax_amount': 2,
                    'invoice_date': '2024-01-02',
                    'invoice_code': 'DUP001',
                    'path': '/b.pdf',
                    'mod_time': '',
                    'size': '',
                },
            ]
            service = InvoiceService(Mock())
            self.assertTrue(service.export_to_excel(pdf_infos, output_path, duplicate_codes={'DUP001'}))
            self.assertTrue(os.path.exists(output_path))

            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active
            # 两行数据都应被高亮（黄色填充）
            for row in range(2, 4):
                self.assertEqual(ws.cell(row=row, column=1).fill.start_color.rgb, '00FFFF00')

    def test_export_to_excel_empty(self):
        """测试空列表导出仍成功创建仅含表头的 Excel"""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, 'out.xlsx')
            service = InvoiceService(Mock())
            self.assertTrue(service.export_to_excel([], output_path))
            self.assertTrue(os.path.exists(output_path))

            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active
            # 空数据时只有表头行和总计行
            self.assertEqual(ws.max_row, 2)


if __name__ == '__main__':
    unittest.main()
